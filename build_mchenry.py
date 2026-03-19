#!/usr/bin/env python3
"""Build McHenry County precinct election data GeoJSON for the interactive map."""

import json
import urllib.request
import openpyxl
from shapely.geometry import shape, mapping, Polygon, MultiPolygon
from shapely.validation import make_valid

# ── Configuration ──────────────────────────────────────────────────────────────
PRECINCT_URL = "https://services1.arcgis.com/6iYC5AXXYapRVNzl/arcgis/rest/services/Precincts/FeatureServer/0/query"
MUNI_URL = "https://services1.arcgis.com/6iYC5AXXYapRVNzl/arcgis/rest/services/Municipal_Boundaries/FeatureServer/0/query"
DETAIL_FILE = "mchenry_data/detail.xlsx"
OUTPUT_FILE = "mchenry_election_data.json"
MUNI_OUTPUT_FILE = "mchenry_municipalities.json"

# DEM-only sheet indices in detail.xlsx
RACE_SHEETS = {
    "race6":  {"sheet_idx": 115, "title": "Senate"},
    "race15": {"sheet_idx": 121, "title": "CD-9"},
    "race16": {"sheet_idx": 123, "title": "CD-10"},
}

ALL_RACE_KEYS = ['race6','race7','race8','race9','race10','race11',
                 'race12','race13','race14','race15','race16','race17']


def download_geojson(base_url, max_records=1000):
    """Download all features from an ArcGIS FeatureServer layer as GeoJSON."""
    count_url = f"{base_url}?where=1%3D1&returnCountOnly=true&f=json"
    resp = urllib.request.urlopen(count_url)
    total = json.loads(resp.read())["count"]
    print(f"  Total features: {total}")

    all_features = []
    offset = 0
    while offset < total:
        url = (f"{base_url}?where=1%3D1&outFields=*&outSR=4326&f=geojson"
               f"&resultRecordCount={max_records}&resultOffset={offset}")
        resp = urllib.request.urlopen(url)
        data = json.loads(resp.read())
        features = data.get("features", [])
        all_features.extend(features)
        print(f"  Downloaded {len(all_features)}/{total}")
        offset += max_records
        if len(features) == 0:
            break

    return {"type": "FeatureCollection", "features": all_features}


def parse_mchenry_sheet(wb, sheet_idx):
    """Parse a McHenry detail.xlsx DEM sheet.

    Format: candidates at cols 3,8,13,18,... (1-indexed), total votes at cols 7,12,17,22,...
    Data rows start at row 4.
    """
    ws = wb[wb.sheetnames[sheet_idx]]

    # Row 2: candidate names at columns 3, 8, 13, ... (1-indexed)
    candidates = []
    col = 3
    while True:
        v = ws.cell(2, col).value
        if v is None:
            break
        name = str(v).strip()
        if name == "WRITE-IN":
            break
        candidates.append(name)
        col += 5

    # Total votes columns: 7, 12, 17, ... (col_start + 4)
    total_cols = [3 + i * 5 + 4 for i in range(len(candidates))]

    results = {}
    for r in range(4, ws.max_row + 1):
        precinct_name = ws.cell(r, 1).value
        if precinct_name is None:
            continue
        precinct_name = str(precinct_name).strip()
        if not precinct_name or precinct_name in ("Total", "Total:"):
            continue

        registered = ws.cell(r, 2).value
        try:
            registered = int(float(registered))
        except (ValueError, TypeError):
            registered = 0

        votes = {}
        total = 0
        for i, cand in enumerate(candidates):
            v = ws.cell(r, total_cols[i]).value
            try:
                v = int(float(v))
            except (ValueError, TypeError):
                v = 0
            votes[cand] = v
            total += v

        # Key is lowercase precinct name for case-insensitive matching
        results[precinct_name.lower()] = {
            "registered": registered,
            "votes": votes,
            "total": total
        }

    return candidates, results


def round_coords(geom):
    """Round geometry coordinates to 5 decimal places."""
    if geom["type"] == "MultiPolygon":
        geom["coordinates"] = [
            [[[round(c[0], 5), round(c[1], 5)] for c in ring]
             for ring in polygon]
            for polygon in geom["coordinates"]
        ]
    elif geom["type"] == "Polygon":
        geom["coordinates"] = [
            [[round(c[0], 5), round(c[1], 5)] for c in ring]
            for ring in geom["coordinates"]
        ]
    return geom


def main():
    # ── Download precinct boundaries ────────────────────────────────────────
    print("Downloading McHenry County precinct boundaries...")
    precincts_gj = download_geojson(PRECINCT_URL)

    # ── Download municipality boundaries ────────────────────────────────────
    print("\nDownloading McHenry County municipality boundaries...")
    munis_gj = download_geojson(MUNI_URL)

    # ── Spatial join: assign municipality to each precinct ──────────────────
    print("\nAssigning municipalities to precincts...")
    muni_shapes = []
    for mf in munis_gj["features"]:
        name = mf["properties"].get("CORPNAME", "")
        if not name:
            continue
        s = shape(mf["geometry"])
        if not s.is_valid:
            s = make_valid(s)
        muni_shapes.append((name.title(), s.buffer(0)))
    print(f"  {len(muni_shapes)} municipality polygons")

    assigned = 0
    for feat in precincts_gj["features"]:
        geom = shape(feat["geometry"])
        centroid = geom.representative_point()
        feat["properties"]["_municipality"] = "Unincorporated"
        for mname, mshape in muni_shapes:
            if mshape.contains(centroid):
                feat["properties"]["_municipality"] = mname
                assigned += 1
                break
    print(f"  Assigned: {assigned}/{len(precincts_gj['features'])} precincts")

    # ── Build precinct name lookup ──────────────────────────────────────────
    # GIS uses "HARTLAND 1", election data uses "Hartland 1" / "McHenry 1"
    # Use case-insensitive lookup
    precinct_map = {}  # lowercase name -> index
    for i, feat in enumerate(precincts_gj["features"]):
        name = feat["properties"]["PrecinctName"]
        precinct_map[name.lower()] = i
        feat["properties"]["_lookup_key"] = name.lower()

    # ── Parse election results ──────────────────────────────────────────────
    print(f"\nParsing {DETAIL_FILE}...")
    wb = openpyxl.load_workbook(DETAIL_FILE, read_only=False)

    all_results = {}
    all_candidates = {}
    for race_key, info in RACE_SHEETS.items():
        print(f"  {race_key} ({info['title']}, sheet idx {info['sheet_idx']})...")
        candidates, results = parse_mchenry_sheet(wb, info["sheet_idx"])
        all_results[race_key] = results
        all_candidates[race_key] = candidates
        print(f"    {len(candidates)} candidates, {len(results)} precincts")
    wb.close()

    # ── Build output features ───────────────────────────────────────────────
    print("\nMerging data...")
    features_out = []
    match_counts = {rk: 0 for rk in RACE_SHEETS}

    for feat in precincts_gj["features"]:
        p = feat["properties"]
        key = p.get("_lookup_key", "")
        muni = p.get("_municipality", "Unincorporated")

        new_props = {
            "name": p["PrecinctName"],
            "municipality": muni,
            "jurisdiction": "mchenry"
        }

        # Initialize all race flags
        for rk in ALL_RACE_KEYS:
            new_props[f"has_{rk}"] = False

        # Merge election results
        for race_key in RACE_SHEETS:
            if key in all_results[race_key]:
                data = all_results[race_key][key]
                new_props[f"has_{race_key}"] = True
                new_props[f"{race_key}_registered"] = data["registered"]
                new_props[f"{race_key}_ballots"] = data["total"]
                new_props[f"{race_key}_total"] = data["total"]

                winner = max(data["votes"], key=data["votes"].get) if data["total"] > 0 else None
                if winner:
                    new_props[f"{race_key}_winner"] = winner
                    new_props[f"{race_key}_winner_pct"] = round(
                        data["votes"][winner] / data["total"] * 100, 1
                    ) if data["total"] > 0 else 0
                    new_props[f"{race_key}_winner_votes"] = data["votes"][winner]

                for cand, votes in data["votes"].items():
                    new_props[f"{race_key}_{cand}"] = votes

                match_counts[race_key] += 1

        features_out.append({
            "type": "Feature",
            "properties": new_props,
            "geometry": round_coords(feat["geometry"])
        })

    print(f"\nMatch results ({len(precincts_gj['features'])} total precincts):")
    for rk in RACE_SHEETS:
        print(f"  {rk}: {match_counts[rk]} matched")

    # Check for unmatched election precincts
    for race_key in RACE_SHEETS:
        for name in all_results[race_key]:
            if name not in precinct_map:
                print(f"  WARNING: {race_key} precinct '{name}' not found in GIS")

    # ── Write election data ─────────────────────────────────────────────────
    output = {"type": "FeatureCollection", "features": features_out}
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, separators=(",", ":"))
    size_mb = len(json.dumps(output, separators=(",", ":"))) / 1024 / 1024
    print(f"\nWritten to {OUTPUT_FILE} ({size_mb:.1f} MB)")

    # ── Write municipality boundaries ───────────────────────────────────────
    muni_features = []
    for mf in munis_gj["features"]:
        name = mf["properties"].get("CORPNAME", "")
        if not name:
            continue
        muni_features.append({
            "type": "Feature",
            "properties": {"name": name.title()},
            "geometry": round_coords(mf["geometry"])
        })

    muni_output = {"type": "FeatureCollection", "features": muni_features}
    with open(MUNI_OUTPUT_FILE, "w") as f:
        json.dump(muni_output, f, separators=(",", ":"))
    print(f"Written to {MUNI_OUTPUT_FILE} ({len(muni_features)} municipalities)")

    # ── Summary ─────────────────────────────────────────────────────────────
    print("\n=== SUMMARY ===")
    for race_key, info in RACE_SHEETS.items():
        candidates = all_candidates[race_key]
        results = all_results[race_key]
        totals = {c: 0 for c in candidates}
        winner_counts = {}
        for data in results.values():
            for c, v in data["votes"].items():
                totals[c] += v
            if data["total"] > 0:
                w = max(data["votes"], key=data["votes"].get)
                winner_counts[w] = winner_counts.get(w, 0) + 1

        grand = sum(totals.values())
        ranked = sorted(candidates, key=lambda c: totals[c], reverse=True)
        print(f"\n{race_key} ({info['title']}):")
        for c in ranked:
            pct = totals[c] / grand * 100 if grand > 0 else 0
            wins = winner_counts.get(c, 0)
            print(f"  {c}: {totals[c]:,} ({pct:.1f}%), {wins} precincts won")


if __name__ == "__main__":
    main()
