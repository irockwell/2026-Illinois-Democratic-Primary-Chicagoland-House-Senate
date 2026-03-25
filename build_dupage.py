#!/usr/bin/env python3
"""Build DuPage County precinct election data GeoJSON for the interactive map."""

import json
import urllib.request
import openpyxl
from shapely.geometry import shape, mapping
from shapely.validation import make_valid

# ── Configuration ──────────────────────────────────────────────────────────────
PRECINCT_URL = "https://gis.dupageco.org/arcgis/rest/services/DuPage_County_IL/Election_Precincts/MapServer/0/query"
MUNI_URL = "https://gis.dupageco.org/arcgis/rest/services/DuPage_County_IL/Municipality/MapServer/0/query"
DETAIL_FILE = "DuPage All Races 3-24-26.xlsx"
OUTPUT_FILE = "dupage_election_data.json"
MUNI_OUTPUT_FILE = "dupage_municipalities.json"

# Sheet numbers for DEM races in detail.xlsx (from Table of Contents)
# Each sheet has: Row 1=race title, Row 2=candidate names (at cols 3,8,13,...), Row 3=headers, Row 4+=data
RACE_SHEETS = {
    "race6":  {"sheet": "2",  "title": "Senate"},
    "race13": {"sheet": "21", "title": "CD-6"},
    "race14": {"sheet": "24", "title": "CD-8"},
}

# Map DuPage ALL CAPS names to Title Case names matching RACE_CONFIG
NAME_MAP = {
    # Senate
    "JULIANA STRATTON": "Juliana Stratton",
    "RAJA KRISHNAMOORTHI": "Raja Krishnamoorthi",
    "ROBIN KELLY": "Robin Kelly",
    "KEVIN RYAN": "Kevin Ryan",
    "SEAN BROWN": "Sean Brown",
    "BRYAN MAXWELL": "Bryan Maxwell",
    "CHRISTOPHER SWANN": "Christopher Swann",
    "AWISI A. BUSTOS": "Awisi A. Bustos",
    "JONATHAN DEAN": "Jonathan Dean",
    "STEVE BOTSFORD JR.": "Steve Botsford Jr.",
    "ADAM DELGADO": "Adam Delgado",
    "Adam Delgado (Write-In)": "Adam Delgado",
    # CD-6
    "SEAN CASTEN": "Sean Casten",
    'JOSEPH "JOEY" RUZEVICH': 'Joseph "Joey" Ruzevich',
    # CD-8
    "NEIL KHOT": "Neil Khot",
    "YASMEEN BANKOLE": "Yasmeen Bankole",
    "KEVIN B. MORRISON": "Kevin B. Morrison",
    "DAN TULLY": "Dan Tully",
    "RYAN VETTICAD": "Ryan Vetticad",
    "MELISSA L. BEAN": "Melissa L. Bean",
    "JUNAID AHMED": "Junaid Ahmed",
    "SANJYOT DUNUNG": "Sanjyot Dunung",
}

def normalize_name(name):
    return NAME_MAP.get(name, name)

ALL_RACE_KEYS = ['race6','race7','race8','race9','race10','race11',
                 'race12','race13','race14','race15','race16','race17']


def download_geojson(base_url, max_records=1000):
    """Download all features from an ArcGIS MapServer layer as GeoJSON."""
    # Get count first
    count_url = f"{base_url}?where=1%3D1&returnCountOnly=true&f=json"
    resp = urllib.request.urlopen(count_url)
    total = json.loads(resp.read())["count"]
    print(f"  Total features: {total}")

    all_features = []
    offset = 0
    while offset < total:
        url = (f"{base_url}?where=1%3D1&outFields=*&outSR=4326&f=geojson"
               f"&resultRecordCount={max_records}&resultOffset={offset}")
        resp = urllib.request.urlopen(url)
        data = json.loads(resp.read())
        features = data.get("features", [])
        all_features.extend(features)
        print(f"  Downloaded {len(all_features)}/{total}")
        offset += max_records
        if len(features) == 0:
            break

    return {"type": "FeatureCollection", "features": all_features}


def parse_dupage_sheet(wb, sheet_name):
    """Parse a DuPage detail.xlsx sheet into {precinct_key: {candidate: votes}}

    precinct_key = "Township Number" e.g. "Addison 1"
    """
    ws = wb[sheet_name]

    # Row 2: candidate names at columns 3, 8, 13, ... (every 5)
    candidates = []
    col = 3
    while True:
        v = ws.cell(2, col).value
        if v is None:
            break
        candidates.append(normalize_name(str(v).strip()))
        col += 5

    # Total Votes column for each candidate: 7, 12, 17, ... (col_start + 4)
    total_cols = [3 + i * 5 + 4 for i in range(len(candidates))]
    # Also: col 2 = Registered Voters, last col = Total (all candidates)

    results = {}
    for r in range(4, ws.max_row + 1):
        precinct_name = ws.cell(r, 1).value
        if precinct_name is None:
            continue
        precinct_name = str(precinct_name).strip()
        if not precinct_name or precinct_name.startswith("Total"):
            continue

        # Handle both old format "Addison 001 - 0137" and new format "Addison 001"
        # Also skip FEDERAL rows like "FEDERAL - F06"
        if precinct_name.startswith("FEDERAL"):
            continue
        twp_prec = precinct_name.split(" - ")[0].strip()  # works for both formats
        # Split into township name and precinct number
        tokens = twp_prec.rsplit(" ", 1)
        if len(tokens) != 2:
            continue
        township = tokens[0]
        try:
            prec_num = int(tokens[1])
        except ValueError:
            continue
        key = f"{township} {prec_num}"

        registered = ws.cell(r, 2).value
        try:
            registered = int(float(registered))
        except (ValueError, TypeError):
            registered = 0

        votes = {}
        total = 0
        for i, cand in enumerate(candidates):
            v = ws.cell(r, total_cols[i]).value
            try:
                v = int(float(v))
            except (ValueError, TypeError):
                v = 0
            votes[cand] = v
            total += v

        # Sum across vote-type rows for the same precinct
        if key in results:
            results[key]["registered"] = max(results[key]["registered"], registered)
            results[key]["total"] += total
            for cand, v in votes.items():
                results[key]["votes"][cand] = results[key]["votes"].get(cand, 0) + v
        else:
            results[key] = {
                "registered": registered,
                "votes": votes,
                "total": total
            }

    return candidates, results


def main():
    # ── Download precinct boundaries ────────────────────────────────────────
    print("Downloading DuPage precinct boundaries...")
    precincts_gj = download_geojson(PRECINCT_URL)

    # ── Download municipality boundaries ────────────────────────────────────
    print("\nDownloading DuPage municipality boundaries...")
    munis_gj = download_geojson(MUNI_URL)

    # ── Spatial join: assign municipality to each precinct ──────────────────
    print("\nAssigning municipalities to precincts...")
    muni_shapes = []
    for mf in munis_gj["features"]:
        name = mf["properties"].get("CITY", "")
        if not name or name == "Uninc":
            continue
        s = shape(mf["geometry"])
        if not s.is_valid:
            s = make_valid(s)
        muni_shapes.append((name, s.buffer(0)))
    print(f"  {len(muni_shapes)} municipality polygons")

    assigned = 0
    for feat in precincts_gj["features"]:
        geom = shape(feat["geometry"])
        centroid = geom.representative_point()
        feat["properties"]["_municipality"] = "Unincorporated"
        for mname, mshape in muni_shapes:
            if mshape.contains(centroid):
                feat["properties"]["_municipality"] = mname
                assigned += 1
                break
    print(f"  Assigned: {assigned}/{len(precincts_gj['features'])} precincts")

    # ── Build precinct name lookup ──────────────────────────────────────────
    precinct_map = {}  # "Township Number" -> index
    for i, feat in enumerate(precincts_gj["features"]):
        p = feat["properties"]
        key = f"{p['TownshipName']} {p['Label']}"
        precinct_map[key] = i

    # ── Parse election results ──────────────────────────────────────────────
    print(f"\nParsing {DETAIL_FILE}...")
    wb = openpyxl.load_workbook(DETAIL_FILE, read_only=False)

    all_results = {}
    all_candidates = {}
    for race_key, info in RACE_SHEETS.items():
        print(f"  {race_key} ({info['title']}, sheet {info['sheet']})...")
        candidates, results = parse_dupage_sheet(wb, info["sheet"])
        all_results[race_key] = results
        all_candidates[race_key] = candidates
        print(f"    {len(candidates)} candidates, {len(results)} precincts")
    wb.close()

    # ── Build output features ───────────────────────────────────────────────
    print("\nMerging data...")
    features_out = []
    match_counts = {rk: 0 for rk in RACE_SHEETS}

    for feat in precincts_gj["features"]:
        p = feat["properties"]
        key = f"{p['TownshipName']} {p['Label']}"
        muni = p.get("_municipality", "Unincorporated")

        new_props = {
            "name": key.upper(),
            "municipality": muni,
            "jurisdiction": "dupage"
        }

        # Initialize all race flags
        for rk in ALL_RACE_KEYS:
            new_props[f"has_{rk}"] = False

        # Merge election results
        for race_key, info in RACE_SHEETS.items():
            if key in all_results[race_key]:
                data = all_results[race_key][key]
                new_props[f"has_{race_key}"] = True
                new_props[f"{race_key}_registered"] = data["registered"]
                new_props[f"{race_key}_ballots"] = data["total"]  # DuPage doesn't separate ballots
                new_props[f"{race_key}_total"] = data["total"]

                winner = max(data["votes"], key=data["votes"].get) if data["total"] > 0 else None
                if winner:
                    new_props[f"{race_key}_winner"] = winner
                    new_props[f"{race_key}_winner_pct"] = round(
                        data["votes"][winner] / data["total"] * 100, 1
                    ) if data["total"] > 0 else 0
                    new_props[f"{race_key}_winner_votes"] = data["votes"][winner]

                for cand, votes in data["votes"].items():
                    new_props[f"{race_key}_{cand}"] = votes

                match_counts[race_key] += 1

        # Round coordinates
        geom = feat["geometry"]
        if geom["type"] == "MultiPolygon":
            geom["coordinates"] = [
                [[[round(c[0], 5), round(c[1], 5)] for c in ring]
                 for ring in polygon]
                for polygon in geom["coordinates"]
            ]
        elif geom["type"] == "Polygon":
            geom["coordinates"] = [
                [[round(c[0], 5), round(c[1], 5)] for c in ring]
                for ring in geom["coordinates"]
            ]

        features_out.append({
            "type": "Feature",
            "properties": new_props,
            "geometry": geom
        })

    print(f"\nMatch results ({len(precincts_gj['features'])} total precincts):")
    for rk in RACE_SHEETS:
        print(f"  {rk}: {match_counts[rk]} matched")

    # ── Write election data ─────────────────────────────────────────────────
    output = {"type": "FeatureCollection", "features": features_out}
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, separators=(",", ":"))
    size_mb = len(json.dumps(output, separators=(",", ":"))) / 1024 / 1024
    print(f"\nWritten to {OUTPUT_FILE} ({size_mb:.1f} MB)")

    # ── Write municipality boundaries ───────────────────────────────────────
    muni_features = []
    for mf in munis_gj["features"]:
        name = mf["properties"].get("CITY", "")
        if not name or name == "Uninc":
            continue
        geom = mf["geometry"]
        if geom["type"] == "MultiPolygon":
            geom["coordinates"] = [
                [[[round(c[0], 5), round(c[1], 5)] for c in ring]
                 for ring in polygon]
                for polygon in geom["coordinates"]
            ]
        elif geom["type"] == "Polygon":
            geom["coordinates"] = [
                [[round(c[0], 5), round(c[1], 5)] for c in ring]
                for ring in geom["coordinates"]
            ]
        muni_features.append({
            "type": "Feature",
            "properties": {"name": name},
            "geometry": geom
        })

    muni_output = {"type": "FeatureCollection", "features": muni_features}
    with open(MUNI_OUTPUT_FILE, "w") as f:
        json.dump(muni_output, f, separators=(",", ":"))
    print(f"Written to {MUNI_OUTPUT_FILE} ({len(muni_features)} municipalities)")

    # ── Summary ─────────────────────────────────────────────────────────────
    print("\n=== SUMMARY ===")
    for race_key, info in RACE_SHEETS.items():
        candidates = all_candidates[race_key]
        results = all_results[race_key]
        totals = {c: 0 for c in candidates}
        winner_counts = {}
        for data in results.values():
            for c, v in data["votes"].items():
                totals[c] += v
            if data["total"] > 0:
                w = max(data["votes"], key=data["votes"].get)
                winner_counts[w] = winner_counts.get(w, 0) + 1

        grand = sum(totals.values())
        ranked = sorted(candidates, key=lambda c: totals[c], reverse=True)
        print(f"\n{race_key} ({info['title']}):")
        for c in ranked:
            pct = totals[c] / grand * 100 if grand > 0 else 0
            wins = winner_counts.get(c, 0)
            print(f"  {c}: {totals[c]:,} ({pct:.1f}%), {wins} precincts won")


if __name__ == "__main__":
    main()
